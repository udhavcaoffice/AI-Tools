import streamlit as st
import pandas as pd
import io
import re
from pdf2image import convert_from_bytes
import pytesseract
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fuzzywuzzy import fuzz

# ==========================================
# 1. PAGE CONFIGURATION & VISUAL OVERHAUL
# ==========================================
st.set_page_config(
    page_title="Udhav Agarwalla & Co | AI Tools",
    page_icon="⚖️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# --- PROFESSIONAL CSS STYLING ---
st.markdown("""
    <style>
        /* 1. GLOBAL FONT & COLOR SETTINGS */
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;800&display=swap');
        
        html, body, [class*="css"] {
            font-family: 'Inter', sans-serif;
            color: #1F2937;
        }

        /* 2. AGGRESSIVE WHITESPACE REMOVAL */
        .block-container {
            padding-top: 0rem !important; /* Removed top padding completely */
            padding-bottom: 2rem !important;
            padding-left: 2rem !important;
            padding-right: 2rem !important;
            max-width: 100% !important;
        }
        
        /* Remove the default top header bar entirely */
        header[data-testid="stHeader"] {
            display: none;
        }

        /* 3. SIDEBAR STYLING */
        [data-testid="stSidebar"] {
            background-color: #F8FAFC; 
            border-right: 1px solid #E2E8F0;
        }
        
        /* 4. CUSTOM HEADER COMPONENT - UPDATED FOR SIZE */
        .custom-header {
            background: linear-gradient(135deg, #0F172A 0%, #1E293B 100%); /* Modern Navy Gradient */
            padding: 40px 30px; /* Increased padding for grander look */
            border-radius: 0px 0px 16px 16px;
            margin-bottom: 30px;
            color: white;
            box-shadow: 0 10px 15px -3px rgba(0, 0, 0, 0.1);
            text-align: center;
        }
        
        /* BIGGER FIRM NAME */
        .custom-header h1 {
            font-size: 42px !important; /* Increased from 26px */
            font-weight: 800 !important; /* Extra Bold */
            margin: 0;
            color: #FFFFFF !important;
            letter-spacing: 1px;
            text-transform: uppercase; /* Makes it look more official */
            line-height: 1.2;
        }
        
        /* BIGGER SUBTITLE */
        .custom-header p {
            font-size: 20px !important; /* Increased from 15px */
            color: #94A3B8;
            margin: 8px 0 0 0;
            font-weight: 500;
            letter-spacing: 0.5px;
        }

        /* 5. "CARD" STYLE FOR TOOLS */
        .tool-card {
            background-color: white;
            padding: 24px;
            border-radius: 12px;
            border: 1px solid #E5E7EB;
            box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05);
            margin-bottom: 20px;
        }
        
        /* 6. BUTTON STYLING */
        .stButton>button {
            background-color: #2563EB; 
            color: white;
            border-radius: 8px;
            border: none;
            padding: 12px 24px;
            font-weight: 600;
            width: 100%;
            transition: all 0.2s;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }
        .stButton>button:hover {
            background-color: #1D4ED8;
            box-shadow: 0 10px 15px -3px rgba(37, 99, 235, 0.3);
            transform: translateY(-1px);
        }

        /* 7. TAB STYLING */
        .stTabs [data-baseweb="tab-list"] {
            gap: 8px;
            border-bottom: 1px solid #E5E7EB;
            margin-bottom: 20px;
        }
        .stTabs [data-baseweb="tab"] {
            height: 50px;
            white-space: pre-wrap;
            background-color: transparent;
            border: none;
            color: #64748B;
            font-weight: 600;
            padding: 0 20px;
        }
        .stTabs [aria-selected="true"] {
            color: #0F172A; /* Darker selected text */
            border-bottom: 3px solid #2563EB;
            background-color: #EFF6FF; /* Subtle blue background for active tab */
        }
        
        /* Hide Footer */
        footer {visibility: hidden;}
    </style>
""", unsafe_allow_html=True)

# ==========================================
# 2. MAIN HEADER
# ==========================================
st.markdown("""
    <div class="custom-header">
        <h1>Udhav Agarwalla & Co.</h1>
        <p>AI-POWERED UTILITIES</p>
    </div>
""", unsafe_allow_html=True)

# ==========================================
# 3. SIDEBAR NAVIGATION
# ==========================================
with st.sidebar:
    st.markdown("### 🛠️ TOOL MENU")
    
    # UPDATED MENU LIST
    selected_tool = st.radio(
        "Select Module:",
        [
            "26AS Automation", 
            "GST Utilities", 
            "Tax Audit Utilities", 
            "Company Audit Utilities"
        ],
        index=0,
        label_visibility="collapsed"
    )
    
    st.markdown("---")
    
    # Contact Card in Sidebar
    st.markdown("""
        <div style="background-color:white; padding:15px; border-radius:8px; border:1px solid #E5E7EB;">
            <small style="color:#6B7280; font-weight:700; letter-spacing:0.5px;">NEED SUPPORT?</small>
            <p style="font-size:13px; margin-top:5px; margin-bottom:0; color:#374151; line-height:1.4;">
                Contact the technical team for support with file formats or errors.
            </p>
        </div>
    """, unsafe_allow_html=True)

# ==========================================
# 4. MAIN APP LOGIC
# ==========================================

if selected_tool == "26AS Automation":
    
    st.markdown("### 📂 26AS Reconciliation Suite")
    st.markdown("Select a function below to process your audit files.")
    
    # TABS
    tab1, tab2, tab3 = st.tabs(["PDF to Excel", "Tally Summary", "Reconciliation"])

    # --- SUB-TOOL 1: PDF to Excel (OCR) ---
    with tab1:
        st.markdown("<div class='tool-card'>", unsafe_allow_html=True) # Start Card
        st.markdown("#### 📄 Convert 26AS PDF to Excel")
        
        col1, col2 = st.columns([1, 1], gap="large")
        
        with col1:
            st.markdown("**Step 1: Upload File**")
            uploaded_pdf = st.file_uploader("Upload 26AS PDF", type="pdf", key="t1", label_visibility="collapsed")
            
            if uploaded_pdf:
                st.markdown("---")
                if st.button("🚀 Start Conversion", key="btn1"):
                    with st.spinner("Scanning PDF... This may take 30-60 seconds..."):
                        try:
                            # 1. Convert PDF to Images
                            images = convert_from_bytes(uploaded_pdf.read())
                            full_text = ""
                            progress_bar = st.progress(0)
                            
                            # 2. Extract Text via OCR
                            for i, image in enumerate(images):
                                text = pytesseract.image_to_string(image, config='--psm 4')
                                full_text += text + "\n"
                                progress_bar.progress((i + 1) / len(images))
                            
                            # 3. Parse Data
                            data = []
                            lines = full_text.split('\n')
                            tan_loose_pattern = re.compile(r'[A-Z]{4}[0-9OIl]{5}[A-Z]')

                            for line in lines:
                                line = line.strip()
                                if len(line) < 15: continue
                                match = tan_loose_pattern.search(line)
                                if match:
                                    tan_code = match.group()
                                    clean_line = re.sub(r'\s+', ' ', line)
                                    parts = clean_line.split(' ')
                                    tan_idx = -1
                                    for idx, part in enumerate(parts):
                                        if tan_code in part:
                                            tan_idx = idx
                                            break
                                    if tan_idx != -1:
                                        name_parts = []
                                        for j in range(tan_idx):
                                            w = parts[j]
                                            if len(w) > 1 and not re.match(r'^\d+$', w) and w.lower() not in ['sr', 'no']:
                                                name_parts.append(w)
                                        party_name = " ".join(name_parts)
                                        party_name = re.sub(r'^[^A-Z]+', '', party_name)

                                        amounts = []
                                        for token in parts[tan_idx+1:]:
                                            token_fix = token.replace('O','0').replace('o','0').replace('l','1').replace('I','1').replace('S','5')
                                            token_clean = re.sub(r'[^\d\.]', '', token_fix)
                                            if re.match(r'^\d+\.?\d{0,2}$', token_clean):
                                                try:
                                                    val = float(token_clean)
                                                    if val > 10:
                                                        amounts.append(val)
                                                except: pass
                                        
                                        final_tax = 0.0
                                        if len(amounts) >= 3:
                                            final_tax = amounts[1]
                                        elif len(amounts) == 2:
                                            final_tax = amounts[1]
                                        elif len(amounts) == 1:
                                            final_tax = amounts[0]
                                        
                                        if len(party_name) > 3 and final_tax > 0:
                                            data.append({
                                                'Name of Party': party_name,
                                                'Amount showing in 26AS': final_tax
                                            })

                            df = pd.DataFrame(data)
                            
                            if not df.empty:
                                df = df.drop_duplicates(subset=['Name of Party'], keep='first')
                                df = df.sort_values('Name of Party').reset_index(drop=True)
                                
                                # 4. Generate Excel
                                output = io.BytesIO()
                                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                    df.to_excel(writer, index=False, sheet_name='26AS Data')
                                    ws = writer.sheets['26AS Data']
                                    ws.column_dimensions['A'].width = 50
                                    ws.column_dimensions['B'].width = 18
                                
                                output.seek(0)
                                st.success(f"✅ Extracted {len(df)} rows successfully.")
                                st.download_button("Download Excel File", data=output, file_name="26AS_Extracted_Data.xlsx")
                            else:
                                st.error("No valid data found. Please check PDF quality.")
                                
                        except Exception as e:
                            st.error(f"Error: {e}")

        with col2:
            st.markdown("""
                <div style="background-color:#F8FAFC; padding:24px; border-radius:12px; border: 1px solid #E2E8F0;">
                    <strong style="color:#0F172A; font-size:16px;">💡 INSTRUCTIONS</strong>
                    <ol style="margin-top:12px; color:#475569; padding-left:20px; line-height:1.6;">
                        <li>Download your 26AS as a <b>PDF</b> from the portal.</li>
                        <li>Drag and drop the file into the box on the left.</li>
                        <li>Click <b>START CONVERSION</b> and wait (OCR is processing).</li>
                        <li>Download the cleaned Excel sheet.</li>
                    </ol>
                </div>
            """, unsafe_allow_html=True)
            
        st.markdown("</div>", unsafe_allow_html=True) # End Card

    # --- SUB-TOOL 2: Tally Summary ---
    with tab2:
        st.markdown("<div class='tool-card'>", unsafe_allow_html=True)
        st.markdown("#### 📒 Tally Ledger to Summary")
        
        uploaded_tally = st.file_uploader("Upload Tally Export (Excel)", type=["xlsx", "xls"], key="t2")
        
        if uploaded_tally:
            if st.button("Generate Summary", key="btn2"):
                try:
                    df = pd.read_excel(uploaded_tally, engine='openpyxl')
                    parties_data = {}
                    
                    # Iterate rows
                    for idx, row in df.iterrows():
                        try:
                            def get_safe(row, idx): return row.iloc[idx] if len(row) > idx else None
                            
                            particulars = str(get_safe(row, 2)).strip() if pd.notna(get_safe(row, 2)) else ''
                            credit = get_safe(row, 5)
                            debit = get_safe(row, 4)
                            
                            if not particulars or particulars == 'nan' or particulars == '': continue
                            if any(skip in particulars for skip in ['Closing Balance', 'nan', '700224', 'Balance', 'Ledger Account', '1-Apr']): continue
                            
                            party = particulars.replace('(Rent)', '').replace('(Interest)', '').strip()
                            if not party or '(as per details)' in party or len(party) < 3 or party.isdigit(): continue
                            
                            amount = None
                            if pd.notna(credit) and credit != '' and credit != 0:
                                try:
                                    c_val = float(credit)
                                    if c_val > 0: amount = c_val
                                except: pass
                            
                            if amount is None and pd.notna(debit) and debit != '' and debit != 0:
                                try:
                                    d_val = float(debit)
                                    if d_val > 0: amount = d_val
                                except: pass
                                
                            if amount and amount > 0:
                                if party in parties_data: parties_data[party] += amount
                                else: parties_data[party] = amount
                        except: pass
                    
                    # Create Excel
                    output = io.BytesIO()
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Party Summary"
                    
                    ws['A1'] = "Name of Party"
                    ws['B1'] = "Amount as per Books"
                    
                    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    
                    for cell in ['A1', 'B1']:
                        ws[cell].fill = header_fill
                        ws[cell].font = header_font
                        ws[cell].border = border
                    
                    row_num = 2
                    total_amount = 0
                    
                    for party in sorted(parties_data.keys()):
                        amt = parties_data[party]
                        total_amount += amt
                        ws[f'A{row_num}'] = party
                        ws[f'B{row_num}'] = amt
                        ws[f'B{row_num}'].number_format = '#,##0.00'
                        ws[f'A{row_num}'].border = border
                        ws[f'B{row_num}'].border = border
                        row_num += 1
                        
                    ws[f'A{row_num}'] = "TOTAL"
                    ws[f'B{row_num}'] = total_amount
                    ws[f'A{row_num}'].font = Font(bold=True)
                    ws[f'B{row_num}'].font = Font(bold=True)
                    ws[f'B{row_num}'].number_format = '#,##0.00'
                    
                    ws.column_dimensions['A'].width = 50
                    ws.column_dimensions['B'].width = 20
                    
                    wb.save(output)
                    output.seek(0)
                    
                    st.success(f"Processed! Found {len(parties_data)} unique parties.")
                    st.download_button("Download Party Summary", data=output, file_name="Party_Summary.xlsx")
                    
                except Exception as e:
                    st.error(f"Error: {e}")
        st.markdown("</div>", unsafe_allow_html=True)

    # --- SUB-TOOL 3: Reconciliation ---
    with tab3:
        st.markdown("<div class='tool-card'>", unsafe_allow_html=True)
        st.markdown("#### 🔍 Reconcile Books vs 26AS")
        
        c1, c2 = st.columns(2)
        with c1:
            file_books = st.file_uploader("Upload Books Summary", type="xlsx", key="f1")
        with c2:
            file_26as = st.file_uploader("Upload 26AS Summary", type="xlsx", key="f2")
            
        st.markdown("<br>", unsafe_allow_html=True)
        threshold = st.slider("Fuzzy Match Sensitivity", 50, 100, 75, help="Lower = looser matching")

        if file_books and file_26as:
            if st.button("Run Reconciliation", key="btn3"):
                try:
                    df_books = pd.read_excel(file_books)
                    df_26as = pd.read_excel(file_26as)
                    
                    df_books.columns = ['Name of Party', 'Amount in Books']
                    df_26as.columns = ['Name of Party', 'Amount in 26AS']
                    
                    df_books = df_books[df_books['Name of Party'].str.upper() != 'TOTAL'].dropna()
                    df_26as = df_26as[df_26as['Name of Party'].str.upper() != 'TOTAL'].dropna()
                    
                    matched_pairs = {}
                    matched_26as_indices = set()
                    
                    # Fuzzy Match Logic
                    for idx_b, party_b in enumerate(df_books['Name of Party']):
                        best_match = None
                        best_score = 0
                        for idx_a, party_a in enumerate(df_26as['Name of Party']):
                            score = fuzz.token_set_ratio(str(party_b).upper(), str(party_a).upper())
                            if score > best_score:
                                best_score = score
                                best_match = idx_a
                        
                        if best_score >= threshold:
                            matched_pairs[idx_b] = best_match
                            matched_26as_indices.add(best_match)

                    reco_data = []
                    
                    # Matches
                    for idx_b, idx_a in matched_pairs.items():
                        b_row = df_books.iloc[idx_b]
                        a_row = df_26as.iloc[idx_a]
                        reco_data.append({
                            'Name of Party': b_row['Name of Party'],
                            'Amount in Books': b_row['Amount in Books'],
                            'Amount in 26AS': a_row['Amount in 26AS'],
                            'Difference': b_row['Amount in Books'] - a_row['Amount in 26AS']
                        })
                    
                    # Only in Books
                    for idx_b in range(len(df_books)):
                        if idx_b not in matched_pairs:
                            row = df_books.iloc[idx_b]
                            reco_data.append({
                                'Name of Party': row['Name of Party'],
                                'Amount in Books': row['Amount in Books'],
                                'Amount in 26AS': 0,
                                'Difference': row['Amount in Books']
                            })
                            
                    # Only in 26AS
                    for idx_a in range(len(df_26as)):
                        if idx_a not in matched_26as_indices:
                            row = df_26as.iloc[idx_a]
                            reco_data.append({
                                'Name of Party': row['Name of Party'],
                                'Amount in Books': 0,
                                'Amount in 26AS': row['Amount in 26AS'],
                                'Difference': -row['Amount in 26AS']
                            })
                            
                    final_df = pd.DataFrame(reco_data).sort_values('Name of Party')
                    
                    total_row = pd.DataFrame({
                        'Name of Party': ['TOTAL'],
                        'Amount in Books': [final_df['Amount in Books'].sum()],
                        'Amount in 26AS': [final_df['Amount in 26AS'].sum()],
                        'Difference': [final_df['Difference'].sum()]
                    })
                    final_df = pd.concat([final_df, total_row], ignore_index=True)
                    
                    # Save to Excel & Format
                    output = io.BytesIO()
                    final_df.to_excel(output, index=False, sheet_name='Reconciliation')
                    output.seek(0)
                    
                    wb = load_workbook(output)
                    ws = wb.active
                    
                    matched_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid') # Green
                    diff_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid') # Red
                    total_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid') # Yellow
                    
                    for row in ws.iter_rows(min_row=2):
                        party = row[0].value
                        diff = row[3].value
                        
                        fill_to_apply = None
                        if party == 'TOTAL': fill_to_apply = total_fill
                        elif diff == 0 or abs(diff) < 1.0: fill_to_apply = matched_fill
                        else: fill_to_apply = diff_fill
                            
                        if fill_to_apply:
                            for cell in row: cell.fill = fill_to_apply
                                
                    final_output = io.BytesIO()
                    wb.save(final_output)
                    final_output.seek(0)
                    
                    st.success("Reconciliation Complete!")
                    st.download_button("Download Reconciliation Statement", data=final_output, file_name="Reconciliation_Statement.xlsx")
                    
                except Exception as e:
                    st.error(f"Error: {e}")
        st.markdown("</div>", unsafe_allow_html=True)

# --- OTHER TOOLS ---
elif selected_tool == "GST Utilities":
    st.markdown("### 📊 GST Utilities")
    st.info("🚧 This module is currently under development.")
    # You can add tabs here later: tab1, tab2 = st.tabs(["GSTR-2A vs Books", "GSTR-1 Analysis"])

elif selected_tool == "Tax Audit Utilities":
    st.markdown("### 📝 Tax Audit Utilities")
    st.info("🚧 This module is currently under development.")
    # Tabs placeholder: tab1, tab2 = st.tabs(["Clause 44 Analysis", "Depreciation Calculator"])

elif selected_tool == "Company Audit Utilities":
    st.markdown("### 🏢 Company Audit Utilities")
    st.info("🚧 This module is currently under development.")
    # Tabs placeholder: tab1, tab2 = st.tabs(["Ratio Analysis", "Schedule III Checks"])
